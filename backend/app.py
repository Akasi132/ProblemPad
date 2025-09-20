import sqlite3
import os
from flask import Flask, g, jsonify, request, send_from_directory
from flask_cors import CORS
from flask import abort, send_file
import io
import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DB_PATH = os.path.join(BASE_DIR, 'reports.db')
XLSX_PATH = os.path.join(BASE_DIR, 'reports.xlsx')

app = Flask(__name__, static_folder=BASE_DIR)
CORS(app)

# Admin token (set via env). If not set, default to 'changeme' for local testing.
ADMIN_TOKEN = os.environ.get('ADMIN_TOKEN', 'changeme')

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DB_PATH)
        db.row_factory = sqlite3.Row
    return db

def init_db():
    db = get_db()
    db.execute('''
    CREATE TABLE IF NOT EXISTS reports (
        id TEXT PRIMARY KEY,
        title TEXT NOT NULL,
        description TEXT NOT NULL,
        severity TEXT,
        solution TEXT,
        created TEXT
    )
    ''')
    db.commit()

@app.teardown_appcontext
def close_connection(exception):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

@app.route('/')
def index():
    # serve the main index from the workspace root
    return send_from_directory(BASE_DIR, 'index.html')

@app.route('/admin.html')
def admin_page():
    token = request.args.get('token')
    if token != ADMIN_TOKEN:
        return abort(403)
    return send_from_directory(BASE_DIR, 'admin.html')

@app.route('/<path:filename>')
def static_files(filename):
    return send_from_directory(BASE_DIR, filename)

@app.route('/api/reports', methods=['GET'])
def list_reports():
    db = get_db()
    cur = db.execute('SELECT * FROM reports ORDER BY created DESC')
    rows = cur.fetchall()
    return jsonify([dict(r) for r in rows])


@app.route('/api/reports.xlsx', methods=['GET'])
def download_reports_xlsx():
    token = request.args.get('token')
    if token != ADMIN_TOKEN:
        return abort(403)
    if not os.path.exists(XLSX_PATH):
        return jsonify({'error':'xlsx not found'}), 404
    return send_file(XLSX_PATH, as_attachment=True, download_name='reports.xlsx')

@app.route('/api/reports', methods=['POST'])
def create_report():
    data = request.get_json() or {}
    # Accept both old and new payload shapes. Require id, title, description, created
    required = ('id','title','description','created')
    if not all(k in data for k in required):
        return jsonify({'error':'missing fields'}), 400
    db = get_db()
    try:
        db.execute('INSERT INTO reports (id,title,description,severity,solution,created) VALUES (?,?,?,?,?,?)',
                   (data['id'], data['title'], data['description'], data.get('severity') or data.get('impact'), data.get('solution'), data['created']))
        db.commit()
    except sqlite3.IntegrityError:
        return jsonify({'error':'id exists'}), 409

    # Append to Excel workbook (create if not exists)
    try:
        if os.path.exists(XLSX_PATH):
            wb = load_workbook(XLSX_PATH)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(['id','startup','startup_desc','title','description','impact','severity','solution','created'])

        row = [
            data.get('id'),
            data.get('startup'),
            data.get('startup_desc'),
            data.get('title'),
            data.get('description'),
            data.get('impact'),
            data.get('severity'),
            data.get('solution'),
            data.get('created')
        ]
        ws.append(row)
        wb.save(XLSX_PATH)
    except Exception as e:
        app.logger.error('xlsx write error: %s', e)
    return jsonify({'ok':True}), 201

@app.route('/api/reports/<id>', methods=['DELETE'])
def delete_report(id):
    db = get_db()
    db.execute('DELETE FROM reports WHERE id = ?', (id,))
    db.commit()
    return jsonify({'ok':True})

if __name__ == '__main__':
    # ensure DB exists
    with app.app_context():
        init_db()
    app.run(host='127.0.0.1', port=8000, debug=True)
